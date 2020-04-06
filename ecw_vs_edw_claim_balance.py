import teradata
import pandas as pd
import sys
import pyodbc
from decimal import Decimal
import openpyxl
import datetime
from openpyxl import load_workbook
from openpyxl.utils import cell
import os

#variables----------------------------------------------------------------------------------------------------
claim = '900335047'
region = '9'

server_ss_region='xrdcwpdbsecw02.hca.corpad.net'
db_ss_region='ecwstage'
login_file=r'C:\Users\QHI6494\Documents\documents_jason\script\python\python_source_file\teradata_login.txt' 
driver_ss='{SQL Server}'
driver_td='Teradata'
authentication_td='LDAP'
mobiledoc_db_suffix='_view'
#---------------------------------------------------------------------------------------------------------------

#information dictionary
dict_info={}


#get region db
conn_string_ss = 'Driver={0};Server={1};Database={2};,Trusted_Connection=yes'.format(
    driver_ss,server_ss_region,db_ss_region)

conn_ss = pyodbc.connect(conn_string_ss)
cursor_ss=conn_ss.cursor() 
query='''
select
    databasename
from ecwstage.dbo.region
where
    regionid = ?
    and activeflag = ?
'''

data_df=pd.read_sql(query,conn_ss,params=(region,1))
if data_df.empty==True:
    print('No region database returned, process cancelled')
    exit()
db_prefix=data_df.iloc[0]['databasename']
db_mobiledoc=db_prefix+mobiledoc_db_suffix


#get mobiledoc invoice
query='''
select
    cast(id as varchar(30)) as claim
    ,cast(encounterid as varchar(30)) as encounter
    ,patientid
    ,invoicedt
    ,servicedt
    ,invoiceamount
    ,balance    
from {0}.dbo.edi_invoice
where id = ?;
'''.format(db_mobiledoc)

mobiledoc_invoice_df=pd.read_sql(query,conn_ss,params=[claim])
if data_df.empty==True:
    print('No mobiledoc claim returned, process cancelled')
    exit()

#mobiledoc encounter
dict_info['mobiledoc_invoice_encounter']=mobiledoc_invoice_df.iloc[0]['encounter']
dict_info['mobiledoc_invoice_claim']=mobiledoc_invoice_df.iloc[0]['claim']




#get mobiledoc charges
query='''
select    
    cast(t.id as varchar(30)) as trxn_id
    ,t.date
    ,t.trtype
    ,t.trrefid
    ,t.trflag    
    ,t.amount
    ,t.description
    ,t.modifieddate
from {0}.dbo.transactions as t
left join {0}.dbo.edi_inv_cpt as cpt
    on t.trrefid = cpt.id    
left join {0}.dbo.edi_invoice as cptinv
    on cpt.invoiceid = cptinv.id    
where
    cptinv.id = ?
    and t.trtype = 'charges'
order by t.id
'''.format(db_mobiledoc)
mobiledoc_charge_df=pd.read_sql(query,conn_ss,params=[claim])



#get mobiledoc payments
query='''
select
    cast(t.id as varchar(30)) as trxn_id
    ,t.date
    ,t.trtype
    ,t.trflag
    ,t.trrefid
    ,-t.amount as amount
    ,t.description
    ,t.modifiedDate    
from {0}.dbo.transactions as t
left join {0}.dbo.edi_paymentdetail as pmtdet
    on t.trrefid = pmtdet.pmtdetailid    
left join {0}.dbo.enc as pmtenc
    on pmtdet.encounterid = pmtenc.encounterid    
left join {0}.dbo.edi_invoice as pmtencinv
    on pmtenc.invoiceid = pmtencinv.id    
where
    pmtencinv.id = ?
    and t.trtype like 'enc%'
    and amount <> 0
order by t.id
'''.format(db_mobiledoc)
mobiledoc_payment_df=pd.read_sql(query,conn_ss,params=[claim])


#get mobiledoc adjustments
query='''
select
    cast(t.id as varchar(30)) as trxn_id
    ,t.date
    ,t.trtype
    ,t.trflag
    ,t.trrefid
    ,-t.amount as amount
    ,t.description
    ,t.modifiedDate    
from {0}.dbo.transactions as t
left join {0}.dbo.edi_inv_adjustments as adj
    on t.trrefid = adj.id
left join {0}.dbo.edi_invoice as adjinv
    on adj.invid = adjinv.id
where
    adjinv.id = ?
    and t.trtype = 'adjustments'
    
union all

select
    cast(t.id as varchar(30)) as trxn_id
    ,t.date
    ,t.trtype
    ,t.trflag
    ,t.trrefid
    ,-t.amount as amount
    ,t.description
    ,t.modifiedDate    
from {0}.dbo.transactions as t
left join {0}.dbo.edi_inv_deleted_adj as adj
    on t.trrefid = adj.refid
left join {0}.dbo.edi_invoice as adjinv
    on adj.invid = adjinv.id
where
    adjinv.id = ?
    and t.trtype = 'adjustments'
'''.format(db_mobiledoc)
mobiledoc_adjustment_df=pd.read_sql(query,conn_ss,params=[claim,claim])

mobiledoc_trxn_df = pd.concat([mobiledoc_charge_df,
                               mobiledoc_payment_df,
                               mobiledoc_adjustment_df])


mobiledoc_claim_balance=mobiledoc_invoice_df.iloc[0]['balance']
mobiledoc_trxn_balance=mobiledoc_trxn_df['amount'].sum()
# if mobiledoc_claim_balance != mobiledoc_trxn_balance:
#     print('mobiledoc transactions do not equal mobiledoc invoice claim balance, process cancelled')
#     sys.exit()
# else:
#     print('mobiledoc transactions equal mobiledoc invoice claim balance')

mobiledoc_trxn_df['encounter']=dict_info['mobiledoc_invoice_encounter']
mobiledoc_trxn_df['claim']=dict_info['mobiledoc_invoice_claim']


# out_file=(r'C:\temp\ecw_claim_test_'+
#           datetime.datetime.now().strftime("%Y%m%d-%H%M%S")+
#           '_.xlsx'
#           )
#   
# mobiledoc_trxn_df.to_excel(out_file,index=False)
# os.startfile(out_file)
# sys.exit()



#td login stuff
source_file=login_file
f=open(source_file, "r")
login_str=str(f.readline())
login_list=login_str.split(sep="|")
host_td, uname_td, pword_td=login_list[0], login_list[1], login_list[2]

#make connection
udaExec=teradata.UdaExec(appName="test", version="1.0", logConsole=False)
with udaExec.connect(method="odbc",
                      system=host_td,
                      username=uname_td, 
                      password=pword_td, 
                      driver=driver_td,
                      authentication=authentication_td) as connect:
    query="""
    select
        claim_dw_id as claim_dw_id
    from edwps_base_views.claim
    where
        source_system_code = '1'
        and claim_src_sys_key = ?        
    """
    edw_claim_df=pd.read_sql(query,connect,params=[claim])        
    edw_claim_dw_id=Decimal(edw_claim_df.iloc[0]['claim_dw_id'])
    
    #create temp table
    query="""
    create multiset volatile table ar_trxn_claim
    (
            relational_method varchar(15) character set latin not casespecific
            ,coid char(5) character set latin not casespecific
            ,claim_dw_id decimal(18,0)
            ,data_server_code char(10) character set latin not casespecific compress ('0001      ','0002      ','0003      ','0004      ','0005      ','0006      ','0007      ')
            ,transaction_dw_id decimal(18,0) not null
            ,transaction_src_sys_key decimal(20,0)
            ,encounter_dw_id decimal(18,0)
            ,transaction_amt decimal(18,3)
            ,transaction_unit_qty decimal(18,3)
            ,transaction_action_num integer compress (1 ,2 ,3 ,4 ,5 ,6 ,7 ,8 ,9 ,10 ,11 ,12 ,13 ,14 )
            ,transaction_type_dw_id decimal(18,0)
            ,transaction_class_dw_id decimal(18,0)
            ,pmt_meth_num integer
            ,entry_date timestamp(0)
            ,transaction_date date format 'yyyy-mm-dd'
            ,patient_dw_id decimal(18,0)
            ,dw_add_date_time timestamp(0)
            ,encounter_src_sys_key decimal(20,0)
            ,claim_src_sys_key decimal(20,0)
    ) primary index (transaction_dw_id)
    on commit preserve rows;    
    """   
    connect.execute(query)
    
    #insert transactions based on claim
    query="""
    insert into ar_trxn_claim
    (
    relational_method
    ,coid
    ,claim_dw_id
    ,data_server_code
    ,transaction_dw_id
    ,transaction_src_sys_key
    ,encounter_dw_id
    ,transaction_amt
    ,transaction_unit_qty
    ,transaction_action_num
    ,transaction_type_dw_id
    ,transaction_class_dw_id
    ,pmt_meth_num
    ,entry_date
    ,transaction_date
    ,patient_dw_id
    ,dw_add_date_time
    ,encounter_src_sys_key
    ,claim_src_sys_key
    )
    select
        'claim' as relational_method
        ,a.coid
        ,a.claim_dw_id
        ,a.data_server_code
        ,a.transaction_dw_id
        ,a.transaction_src_sys_key
        ,a.encounter_dw_id
        ,a.transaction_amt
        ,a.transaction_unit_qty
        ,a.transaction_action_num
        ,a.transaction_type_dw_id
        ,a.transaction_class_dw_id
        ,a.pmt_meth_num
        ,a.entry_date
        ,a.transaction_date
        ,a.patient_dw_id
        ,a.dw_add_date_time
        ,(select encounter_src_sys_key from edwps_base_views.encounter e where e.encounter_dw_id=a.encounter_dw_id and source_system_code = a.source_system_code) as encounter_src_sys_key
        ,(select claim_src_sys_key from edwps_base_views.claim clm where clm.claim_dw_id= a.claim_dw_id and clm.source_system_code = a.source_system_code) as claim_src_sys_key
    from edwps_base_views.ar_transaction a
    where
        a.claim_dw_id = ?
        and a.source_system_code = '1'        
        and a.transaction_class_dw_id in (1, 2, 6, 7, 8) --claim level
        and a.transaction_amt <> 0;      
    """
    connect.execute(query, (edw_claim_dw_id,) )
        
    #insert transactions based on encounter
    query="""
    insert into ar_trxn_claim
    (
    relational_method
    ,coid
    ,claim_dw_id
    ,data_server_code
    ,transaction_dw_id
    ,transaction_src_sys_key
    ,encounter_dw_id
    ,transaction_amt
    ,transaction_unit_qty
    ,transaction_action_num
    ,transaction_type_dw_id
    ,transaction_class_dw_id
    ,pmt_meth_num
    ,entry_date
    ,transaction_date
    ,patient_dw_id
    ,dw_add_date_time
    ,encounter_src_sys_key
    ,claim_src_sys_key
    )    
    select distinct
        'encounter' as relational_method
        ,ar.coid
        ,ar.claim_dw_id
        ,ar.data_server_code
        ,ar.transaction_dw_id
        ,ar.transaction_src_sys_key
        ,ar.encounter_dw_id
        ,ar.transaction_amt as transaction_amt
        ,ar.transaction_unit_qty
        ,ar.transaction_action_num
        ,ar.transaction_type_dw_id
        ,ar.transaction_class_dw_id
        ,ar.pmt_meth_num
        ,ar.entry_date
        ,ar.transaction_date
        ,ar.patient_dw_id
        ,ar.dw_add_date_time
        ,(select encounter_src_sys_key from edwps_base_views.encounter e where e.encounter_dw_id=ar.encounter_dw_id and e.source_system_code = ar.source_system_code) as encounter_src_sys_key
        ,(select claim_src_sys_key from edwps_base_views.claim clm where clm.claim_dw_id= ar.claim_dw_id and clm.source_system_code = ar.source_system_code) as claim_src_sys_key        
    from edwps_base_views.ar_transaction ar    
    inner join ar_trxn_claim arc
        on ar.encounter_dw_id = arc.encounter_dw_id
    where
        ar.source_system_code = '1';
    """    
    connect.execute(query)
    
    #retrieve edw ar trxn data 
    query="""
    select
        relational_method
        ,coid
        ,otranslate(cast(claim_dw_id as varchar(30)),'.','') as claim_dw_id
        ,data_server_code
        ,otranslate(cast(transaction_dw_id as varchar(30)),'.','') as transaction_dw_id
        ,otranslate(cast(transaction_src_sys_key as varchar(30)),'.','') as transaction_src_sys_key
        ,otranslate(cast(encounter_dw_id as varchar(30)),'.','') as encounter_dw_id
        ,transaction_amt as transaction_amt
        ,transaction_unit_qty
        ,transaction_action_num
        ,transaction_type_dw_id
        ,otranslate(cast(transaction_class_dw_id as varchar(5)),'.','') as transaction_class_dw_id
        ,pmt_meth_num
        ,entry_date
        ,transaction_date
        ,otranslate(cast(patient_dw_id as varchar(30)),'.','') as patient_dw_id
        ,dw_add_date_time
        ,otranslate(cast(encounter_src_sys_key as varchar(30)),'.','') as encounter_src_sys_key
        ,otranslate(cast(claim_src_sys_key as varchar(30)),'.','') as claim_src_sys_key
    from ar_trxn_claim    
    """
    edw_ar_trxn_df=pd.read_sql(query,connect)

ar_claim_bal_claim_df=edw_ar_trxn_df.loc[edw_ar_trxn_df['relational_method']=='claim']
ar_claim_bal_encounter_df=edw_ar_trxn_df.loc[edw_ar_trxn_df['relational_method']=='encounter']
ar_claim_bal_claim=ar_claim_bal_claim_df['transaction_amt'].sum()
ar_claim_bal_encounter=ar_claim_bal_encounter_df['transaction_amt'].sum()
ar_claim_bal_claim_var=ar_claim_bal_claim-mobiledoc_claim_balance
ar_claim_bal_encounter_var=ar_claim_bal_encounter-mobiledoc_claim_balance

# out_file=(r'C:\temp\ecw_claim_test_'+
#           datetime.datetime.now().strftime("%Y%m%d-%H%M%S")+
#           '_.xlsx'
#           )
#   
# ar_claim_bal_encounter_df.to_excel(out_file,index=False)
# os.startfile(out_file)
# sys.exit()

        


edw_ar_1_df=pd.merge(
    mobiledoc_trxn_df[
        ['trxn_id','trtype','trflag','amount','encounter','claim']
        ],
    ar_claim_bal_claim_df[
        ['transaction_src_sys_key','transaction_class_dw_id','transaction_amt','encounter_src_sys_key','encounter_dw_id','claim_src_sys_key','claim_dw_id']
        ],
    how='left',
    left_on='trxn_id',
    right_on='transaction_src_sys_key')

# out_file=(r'C:\temp\ecw_claim_test_'+
#           datetime.datetime.now().strftime("%Y%m%d-%H%M%S")+
#           '_.xlsx'
#           )
#   
# edw_ar_1_df.to_excel(out_file,index=False)
# os.startfile(out_file)
# sys.exit()

edw_ar_2_df=pd.merge(
    edw_ar_1_df,
    ar_claim_bal_encounter_df[
        ['transaction_src_sys_key','transaction_class_dw_id','transaction_amt','encounter_src_sys_key','encounter_dw_id','claim_src_sys_key','claim_dw_id']
        ],
    how='left',
    left_on='trxn_id',
    right_on='transaction_src_sys_key')


# out_file=(r'C:\temp\ecw_claim_test_'+
#           datetime.datetime.now().strftime("%Y%m%d-%H%M%S")+
#           '_.xlsx'
#           )
#     
# edw_ar_2_df.to_excel(out_file,index=False)
# os.startfile(out_file)
# sys.exit()


edw_ar_2_df.columns=['transaction_ecw','trtype_ecw','trflag_ecw','amount_ecw','encounter_ecw', 'claim_ecw',
                     'transaction_edw_ar_claim','transaction_class_edw_ar_claim','amount_edw_ar_claim','encounter_edw_ar_claim','encounter_dw_id_edw_ar_claim','claim_edw_ar_claim','claim_dw_id_edw_ar_claim',
                     'transaction_edw_ar_enc','transaction_class_edw_ar_enc','amount_edw_ar_enc','encounter_edw_ar_enc','encounter_dw_id_edw_ar_enc','claim_edw_ar_enc','claim_dw_id_edw_ar_enc']


# out_file=(r'C:\temp\ecw_claim_test_'+
#           datetime.datetime.now().strftime("%Y%m%d-%H%M%S")+
#           '_.xlsx'
#           )
#    
# edw_ar_2_df.to_excel(out_file,index=False)
# os.startfile(out_file)
# sys.exit()


edw_ar_3_df=edw_ar_2_df[[
    'transaction_ecw'
    ,'transaction_edw_ar_claim'
    ,'transaction_edw_ar_enc'
    ,'trtype_ecw'
    ,'trflag_ecw'
    ,'transaction_class_edw_ar_claim'
    ,'transaction_class_edw_ar_enc'
    ,'encounter_ecw'
    ,'encounter_edw_ar_claim'
    ,'encounter_edw_ar_enc'
    ,'encounter_dw_id_edw_ar_claim'
    ,'encounter_dw_id_edw_ar_enc'
    ,'claim_ecw'
    ,'claim_edw_ar_claim'
    ,'claim_edw_ar_enc'    
    ,'claim_dw_id_edw_ar_claim'
    ,'claim_dw_id_edw_ar_enc'
    ,'amount_ecw'
    ,'amount_edw_ar_claim'
    ,'amount_edw_ar_enc'    
    ]]
    


out_file=(r'C:\temp\ecw_claim_analysis_'+
          datetime.datetime.now().strftime("%Y%m%d-%H%M%S")+
          '_.xlsx'
          )

edw_ar_3_df.to_excel(out_file,index=False)
wb=load_workbook(filename=out_file, read_only=False)

ws=wb.active
last_row=ws.max_row
last_col=ws.max_column
total_row=last_row+2

rows=ws.iter_rows(min_row=1,max_row=1)
r1=next(rows)
headings=[cell.get_column_letter(c.column) for c in r1 if 'amount' in c.value]

for h in headings:
    sum_range=h+str(total_row)
    sum_formula='=sum({0}{1}:{0}{2})'.format(h,2,last_row)
    ws[sum_range]=sum_formula
    for r in range(1,total_row+1):
        temp_range=h+str(r)
        ws[temp_range].number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'

for c in range(1,last_col+1):
    ws.column_dimensions[cell.get_column_letter(c)].width=30
    
wb.save(out_file)    
wb.close()

print('Success')
os.startfile(out_file)
              
        
